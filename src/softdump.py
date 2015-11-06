from datetime import date

import openpyxl
from openpyxl.cell import get_column_letter 
SoftList = '../docs/SoftList.rst'

today = date.today()
today = today.strftime("%A %d %B %Y")

head = """
Current Lab Software List
#########################

As of: %s

..  csv-table::
    :header: Software

""" % today

fout = open(SoftList,'wb')
fout.write(head)

wb = openpyxl.load_workbook('../data/COIS Installed Software.xlsx')
type(wb)
sheets = wb.get_sheet_names()
print sheets
sheet = wb.get_sheet_by_name('Sheet1')
print sheet['A3'].value

rows = sheet.max_row+1
print sheet.max_column

row1 = get_column_letter(1)
for item_num in range(5,rows):
    cn = row1 + str(item_num)
    print sheet[cn].value
    fout.write("    %s\n" % sheet[cn].value)

fout.close()
